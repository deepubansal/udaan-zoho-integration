from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle
from datetime import datetime
date_style = NamedStyle(name='datetime', number_format='DD/MM/YYYY')

payments = []


class Payment(object):

    def __init__(self, Payment_Number_Prefix, Payment_Number_Suffix, Payment_Type, Date, Amount, Description, Reference_Number, Invoice_Amount, Invoice_Number, Deposit_To):
        self.Payment_Number_Prefix = Payment_Number_Prefix
        self.Payment_Number_Suffix = Payment_Number_Suffix
        self.Payment_Type = Payment_Type
        self.Date = Date
        self.Amount = Amount
        self.Description = Description
        self.Reference_Number = Reference_Number
        self.Invoice_Amount = Invoice_Amount
        self.Invoice_Number = Invoice_Number
        self.Deposit_To = Deposit_To

    def get_row(self):
        row = [self.Payment_Number_Prefix, self.Payment_Number_Suffix, self.Payment_Type, self.Date, self.Amount,
               self.Description, self.Reference_Number, self.Invoice_Amount, self.Invoice_Number, self.Deposit_To]
        return row


class PaymentSettlement(object):

    def __init__(self, Invoice_Id, Invoice_Date, Description, Settled_For, Buyer_Phone_Number, PaymentReferenceId, Payment_Date, PaymentAmount, Adjusted_Value, Settled_Against, Entry_Type):
        self.Invoice_Id = Invoice_Id
        self.Invoice_Date = datetime.strptime(Invoice_Date, '%d %b %y')
        self.Description = Description
        self.Order_Id = Settled_For
        self.Buyer_Phone_Number = Buyer_Phone_Number
        self.PaymentReferenceId = PaymentReferenceId
        self.Payment_Date = self.read_date(Payment_Date)
        self.PaymentAmount = PaymentAmount
        self.Adjusted_Value = Adjusted_Value

    def read_date(self, d_string):
        try:
            d = datetime.strptime(d_string, '%d-%b-%Y')
        except ValueError:
            d = datetime.strptime(d_string, '%d %b %y')
        return d

    def __str__(self):
        return ('{}\t{}\t{}\t AdjustedValue: {}'.format(self.Invoice_Id, self.Description, self.PaymentReferenceId, self.Adjusted_Value))


class BankTransfer(object):

    def __init__(self, UTR, Amount, Due_Date, Payment_Date, To_Ac_Number):
        self.UTR = UTR
        self.Amount = Amount
        self.Due_Date = datetime.strptime(Due_Date, '%d %b %y')
        self.Payment_Date = datetime.strptime(Payment_Date, '%d %b %y')
        self.To_Ac_Number = To_Ac_Number

    def __str__(self):
        return '{}\t{}\t{}\t{}\t{}'.format(self.UTR, self.Amount, self.Due_Date, self.Payment_Date, self.To_Ac_Number)


class RTO(object):

    def __init__(self, Invoice_Id, Invoice_Date, Description, Order_Id, Amount, RTO_date):
        self.Invoice_Id = Invoice_Id
        self.Invoice_Date = datetime.strptime(Invoice_Date, '%d %b %y')
        self.Description = Description
        self.Order_Id = Order_Id
        self.Amount = Amount
        self.RTO_Date = datetime.strptime(RTO_date, '%d %b %y')

    def __str__(self):
        return '{}\t{}\t Amount: {}'.format(self.Invoice_Id, self.Description, self.Amount)


class Reversal(object):

    def __init__(self, Reversal_Date, Description, PaymentReferenceId, Payment_Date, PaymentAmount, Value_Adjusted_Value, Invoice_Data_Download):
        self.Reversal_Date = datetime.strptime(Reversal_Date, '%d %b %y')
        self.Description = Description
        self.PaymentReferenceId = PaymentReferenceId
        self.Payment_Date = datetime.strptime(Payment_Date, '%d %b %y')
        self.PaymentAmount = PaymentAmount
        self.Adjusted_Value = Value_Adjusted_Value
        self.Invoice_Data_Download = Invoice_Data_Download

    def __str__(self):
        return '{}\t{}\t{}\t AdjustedValue: {}'.format(self.Reversal_Date, self.PaymentReferenceId, self.Description, self.Adjusted_Value)


class ApprovedRefunds(object):

    def __init__(self, For_Invoice_Id, Invoice_Date, Description, Order_Id, Return_Id, Invoice_Amount, Refund_Amount, Platform_Charges_Refunded, Total_Amount_Deductible, Platform_Charges_Refund_Credit_Note, Refund_Date):
        self.For_Invoice_Id = For_Invoice_Id
        self.Invoice_Date = datetime.strptime(Invoice_Date, '%d %b %y')
        self.Description = Description
        self.Order_Id = Order_Id
        self.Return_Id = Return_Id
        self.Invoice_Amount = Invoice_Amount
        self.Refund_Amount = Refund_Amount
        self.Refund_Date = datetime.strptime(Refund_Date, '%d %b %y')

    def __str__(self):
        return '{}\t{}\t{}\t RefundAmount: {}'.format(self.For_Invoice_Id, self.Invoice_Amount, self.Description, self.Refund_Amount)


class Charges(object):

    def __init__(self, Invoice_Date, Invoice_Number, Description, Invoice_Amount, Invoice_Download_Link):
        self.Invoice_Date = datetime.strptime(Invoice_Date, '%d %b %y')
        self.Invoice_Number = Invoice_Number
        self.Description = Description
        self.Invoice_Amount = float(Invoice_Amount.replace(',', ''))
        self.Invoice_Download_Link = Invoice_Download_Link

    def __str__(self):
        return '{}\t{}\t{}\t Amount: {}'.format(self.Invoice_Date, self.Invoice_Number, self.Description, self.Invoice_Amount)


class ZohoPayments(object):

    def __init__(self, Payment_Number, Payment_Number_Prefix, Payment_Number_Suffix, CustomerPayment_ID, Date, CustomerID, Customer_Name, Tax_Type, Payment_Type, Mode, Description, Exchange_Rate, Amount, Deposit_To, Bank_Charges, Reference_Number, InvoicePayment_ID, Invoice_Number, Invoice_Amount, Withholding_Tax_Amount, Unused_Amount, Tax_Account, Currency_Code, Created_Time):
        self.Payment_Number = Payment_Number
        self.Payment_Number_Prefix = Payment_Number_Prefix
        self.Payment_Number_Suffix = Payment_Number_Suffix
        self.CustomerPayment_ID = CustomerPayment_ID
        self.Date = Date
        self.CustomerID = CustomerID
        self.Customer_Name = Customer_Name
        self.Tax_Type = Tax_Type
        self.Payment_Type = Payment_Type
        self.Mode = Mode
        self.Description = Description
        self.Exchange_Rate = Exchange_Rate
        self.Amount = Amount
        self.Deposit_To = Deposit_To
        self.Bank_Charges = Bank_Charges
        self.Reference_Number = Reference_Number
        self.InvoicePayment_ID = InvoicePayment_ID
        self.Invoice_Number = Invoice_Number
        self.Invoice_Amount = Invoice_Amount
        self.Withholding_Tax_Amount = Withholding_Tax_Amount
        self.Unused_Amount = Unused_Amount
        self.Tax_Account = Tax_Account
        self.Currency_Code = Currency_Code
        self.Created_Time = Created_Time

    def __str__(self):
        return '{}\t{}\t{}\t Amount: {}'.format(self.Date, self.Invoice_Number, self.Description, self.Invoice_Amount)


def read_zoho_payments_sheet(file):
    wb = load_workbook(file)
    ws = wb['Sheet1']
    header = [cell.value for cell in ws[1]]
    payment_settlements = []
    for row in ws.iter_rows(row_offset=1):
        if not row[0].value:
            break
        values = {}
        for key, cell in zip(header, row):
            values[key.replace(' ', '_')] = cell.value
        payment_settlement = ZohoPayments(**values)
        payment_settlements.append(payment_settlement)
    return payment_settlements


def read_settlement_files(paths):
    settlements = []
    for path in paths:
        settlements.extend(read_settlements(path))
    return settlements


def read_settlements(settlement_file_path):
    wb = load_workbook(settlement_file_path)
    ws = wb['Settlements']
    header = [cell.value for cell in ws[1]]
    payment_settlements = []
    for row in ws.iter_rows(row_offset=1):
        if not row[0].value:
            break
        values = {}
        for key, cell in zip(header, row):
            values[key.replace(' ', '_')] = cell.value
        payment_settlement = PaymentSettlement(**values)
        payment_settlements.append(payment_settlement)
    return payment_settlements


def read_rto_sheet(file_path):
    wb = load_workbook(file_path)
    ws = wb['RTO']
    header = [cell.value for cell in ws[1]]
    rtos = []
    for row in ws.iter_rows(row_offset=1):
        if not row[0].value:
            break
        values = {}
        for key, cell in zip(header, row):
            values[key.replace(' ', '_')] = cell.value
        rto = RTO(**values)
        rtos.append(rto)
    return rtos


def read_reversal_sheet(file_path):
    wb = load_workbook(file_path)
    ws = wb['Reversals Credited']
    header = [cell.value for cell in ws[1]]
    rtos = []
    for row in ws.iter_rows(row_offset=1):
        if not row[0].value:
            break
        values = {}
        for key, cell in zip(header, row):
            values[key.replace(' ', '_').replace('/', '_')] = cell.value
        rto = Reversal(**values)
        rtos.append(rto)
    return rtos


def read_payments_sheet(file_path):
    wb = load_workbook(file_path)
    ws = wb['All Payments During Period']
    header = [cell.value for cell in ws[1]]
    rtos = []
    for row in ws.iter_rows(row_offset=1):
        if not row[0].value:
            break
        values = {}
        for key, cell in zip(header, row):
            values[key.replace(' ', '_').replace('/', '')] = cell.value
        rto = BankTransfer(**values)
        rtos.append(rto)
    return rtos


def read_approved_refund_sheet(file_path):
    wb = load_workbook(file_path)
    ws = wb['All Approved Refunds']
    header = [cell.value for cell in ws[1]]
    rtos = []
    for row in ws.iter_rows(row_offset=1):
        if not row[0].value:
            break
        values = {}
        for key, cell in zip(header, row):
            values[key.replace(' ', '_').replace('/', '')] = cell.value
        rto = ApprovedRefunds(**values)
        rtos.append(rto)
    return rtos


def read_platform_charges_sheet(file_path):
    wb = load_workbook(file_path)
    ws = wb['All Platform Charges']
    return read_charges(ws)


def read_logistics_charges_sheet(file_path):
    wb = load_workbook(file_path)
    ws = wb['All Logistics Charges']
    return read_charges(ws)


def read_advertisement_charges_sheet(file_path):
    wb = load_workbook(file_path)
    ws = wb['All Advertisement Charges']
    return read_charges(ws)


def read_charges(ws):
    header = [cell.value for cell in ws[1]]
    charges = []
    for row in ws.iter_rows(row_offset=1):
        if not row[0].value:
            break
        values = {}
        for key, cell in zip(header, row):
            key = 'Invoice Download Link' if key == 'Download Link' else key
            values[key.replace(' ', '_').replace('/', '')] = cell.value
        charge = Charges(**values)
        charges.append(charge)
    return charges


def read_other_charges(file_path):
    wb = load_workbook(file_path)
    ws = wb['All Other Charges']
    charges = []
    for row in ws.iter_rows(row_offset=1):
        if not row[0].value:
            break
        import re
        print(row[1].value)
        matched = re.match('(Fulfilment/Packaging Material ordered on )([0-9]{2}-[0-9]{2}-[0-9]{4}) vide order-id ([A-Z0-9]{14})',
                           row[1].value)
        if matched:
            groups = matched.groups()
            values = {'Invoice_Date': row[0].value,
                    'Invoice_Number': groups[2],
                    'Description': row[1].value,
                    'Invoice_Amount': str(row[2].value),
                    'Invoice_Download_Link': ''
                    }
            charge = Charges(**values)
            charges.append(charge)
    return charges


def write_payments(payment_file_path, payments):
    wb = Workbook()
    excel_header = ["Payment Number Prefix", "Payment Number Suffix", "Payment Type", "Date", "Amount",
                    "Description", "Reference Number", "Invoice Amount", "Invoice Number", "Deposit To", "Over"]
    sheet = wb.active
    sheet.append(excel_header)
    for i, payment in enumerate(payments):
        row = payment.get_row()
        row.append(True if abs(payment.Amount - payment.Invoice_Amount) > 2 else False)
        sheet.append(row)
        sheet['D{}'.format(i + 1)].style = date_style
    wb.save(payment_file_path)
